#!/usr/bin/python3
""" 
   Author        :     Amir W. Fathy
   Date          :     23 sep 2023
   Description   :     simple program to parse functions into excel sheet
"""
import os
import openpyxl

folder_path = os.getcwd()
files_with_h_extension = []

for filename in os.listdir(folder_path):
    if filename.endswith(".h"):
        files_with_h_extension.append(filename)

print("Files with '.h' extension in the folder:")
for filename in files_with_h_extension:
    print(filename)

excel_file_path = "functions.xlsx"

workbook = openpyxl.Workbook()


uniqueID = 1

for i in files_with_h_extension:
    with open(i, "r") as file:
        worksheet = workbook.create_sheet(title=i)
        functionNames = file.readlines()
        for index, item in enumerate(functionNames, start=2):

            worksheet.cell(row=1, column=1, value="function name")
            worksheet.cell(row=1, column=2, value="Unique ID")

            worksheet.cell(row=index, column=1, value=item.removesuffix("\n"))
            worksheet.cell(row=index, column=2,
                           value=f'0x{uniqueID:03X}')
            uniqueID += 1

del workbook['Sheet']
workbook.save(excel_file_path)

print("File names with '.h' extension and unique IDs appended to the Excel file successfully.")
